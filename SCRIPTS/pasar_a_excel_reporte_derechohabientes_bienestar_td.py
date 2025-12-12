import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from copy import copy

# 📁 Rutas
carpeta_csv = "/Users/Arturo/AGRICULTURA/FERTILIZANTES/TABLAS_DINAMICAS/"
ruta_encabezado = "/Users/Arturo/AGRICULTURA/INFORMES/INFORMES BIENESTAR/formato_Fertilizantes - seguimiento beneficiarios.xlsx"

# 🧠 Copiar valores y formato de celdas
def copiar_celda_origen_a_destino(celda_origen, celda_destino):
    celda_destino.value = celda_origen.value
    if celda_origen.has_style:
        celda_destino.font = copy(celda_origen.font)
        celda_destino.border = copy(celda_origen.border)
        celda_destino.fill = copy(celda_origen.fill)
        celda_destino.number_format = copy(celda_origen.number_format)
        celda_destino.protection = copy(celda_origen.protection)
        celda_destino.alignment = copy(celda_origen.alignment)

# 🔁 Procesar cada archivo CSV
for archivo in os.listdir(carpeta_csv):
    if archivo.endswith(".csv") and "Fertilizantes - seguimiento beneficiarios -" in archivo:
        ruta_csv = os.path.join(carpeta_csv, archivo)
        df = pd.read_csv(ruta_csv, dtype=str)

        # 📘 Crear nuevo archivo Excel
        wb_nuevo = Workbook()
        ws_nuevo = wb_nuevo.active

        # 📖 Cargar encabezado con formato
        wb_formato = load_workbook(ruta_encabezado)
        ws_formato = wb_formato.active

        # Copiar celdas A1:L2 con estilos
        for fila in range(1, 3):  # filas 1 y 2
            for col in range(1, 13):  # columnas A a L
                celda_origen = ws_formato.cell(row=fila, column=col)
                celda_destino = ws_nuevo.cell(row=fila, column=col)
                copiar_celda_origen_a_destino(celda_origen, celda_destino)

        # Copiar celdas combinadas (merge)
        for rango in ws_formato.merged_cells.ranges:
            if str(rango).startswith("A1") or str(rango).startswith("B1") or str(rango).startswith("C1") or "L2" in str(rango):
                ws_nuevo.merge_cells(str(rango))

        # Copiar contenido CSV desde la fila 3
        for i, row in df.iterrows():
            for j, value in enumerate(row.tolist(), start=1):
                # Si la columna es 'F' o 'G', convertir a número decimal si es posible
                col_letra = get_column_letter(j)
                if col_letra in ['F', 'G']:
                    try:
                        ws_nuevo.cell(row=i + 3, column=j, value=float(value))
                    except:
                        ws_nuevo.cell(row=i + 3, column=j, value=value)  # deja el texto si no es número
                else:
                    ws_nuevo.cell(row=i + 3, column=j, value=value)

        # Ajustar ancho de columnas
        for col in ws_nuevo.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws_nuevo.column_dimensions[col_letter].width = max_length + 2

        # Guardar con mismo nombre pero como .xlsx
        ruta_xlsx = os.path.join(carpeta_csv, archivo.replace(".csv", ".xlsx"))
        wb_nuevo.save(ruta_xlsx)

        print(f"✅ Archivo procesado con formato: {archivo.replace('.csv', '.xlsx')}")

print("🎉 Todos los archivos se procesaron con encabezado formateado correctamente.")